#!/usr/bin/env python3
"""
Social Security deferment + break-even calculator.

Original goal:
    Estimate year-by-year Social Security retirement payments when a person
    starts benefits at a chosen age and continues working at a constant salary.

Phase 1 (break-even, added per Aug-2026 discussion):
    - Pairwise crossover ages between claim-age scenarios (matrix + long table).
    - Lifespan sensitivity: cumulative benefits paid at checkpoint ages.
    - Plain-English recommendation based on the input's expected lifespan.

Phase 2 (input flexibility):
    - `annual_salary` accepts either a scalar OR an object keyed by attained
      age. Missing ages carry the previous value forward (last-observation
      carried forward); ages below the earliest listed key are treated as $0.
    - `benefits_by_claim_age` (optional): map whole-year claim age to the
      monthly benefit shown on the user's SSA statement. When present, the
      script uses those numbers verbatim as the starting benefit at that age.
      FRA recomputation still uses the PIA-based rule (see caveat below).

Phase 3:
    - `lifespan_checkpoints` for a sensitivity view (default 70, 75, 80, 85,
      90, 95). The expected end-of-life age is always included.

Phase 4:
    - Writes an .xlsx workbook (tabs: Summary, Break-Even, Sensitivity,
      Expected Value, Projection). Requires `openpyxl` (pip install openpyxl).
      If openpyxl is missing, the CSV output still runs and the script prints
      install help.

Phase 5 (actuarial expected value):
    - Uses SSA 2021 Period Life Table q(x) values to weight each possible
      death age by its probability, producing an expected lifetime benefit
      for each claim age. This removes the need to guess a single lifespan.
    - Optional input `"sex"`: "male" or "female" (default "male"). Selects
      which mortality table to use.
    - Source: https://www.ssa.gov/oact/STATS/table4c6.html

This is intentionally a simple, auditable model:
    - It uses the SSA-provided full-retirement-age monthly benefit as the base.
    - It applies SSA's early/late claiming percentage rules.
    - It applies SSA's retirement earnings test while the worker is below FRA.
    - It tracks benefits withheld because of work as deferred, not lost.
    - It does NOT adjust for COLA, inflation, taxation, Medicare premiums,
      future benefit recomputation from new earnings, or NPV.
    - It does NOT model SSA's special first-year monthly earnings rule.

Primary SSA evidence used in this script:
    1. Early/delayed claiming percentages:
       https://www.ssa.gov/OACT/ProgData/ar_drc.html
    2. Early retirement monthly reduction formula:
       https://www.ssa.gov/benefits/retirement/planner/agereduction.html
    3. Delayed retirement credits:
       https://www.ssa.gov/benefits/retirement/planner/delayret.html
    4. Working while receiving benefits and the earnings test:
       https://www.ssa.gov/benefits/retirement/planner/whileworking.html
    5. Annual earnings-test exempt amounts:
       https://www.ssa.gov/OACT/COLA/rtea.html
"""

from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any


MONTHS_PER_YEAR = 12
DEFAULT_LIFESPAN_CHECKPOINTS = [70, 75, 80, 85, 90, 95]
ACTUARIAL_MAX_AGE = 100

# SSA 2021 Period Life Table — q(x) = probability of dying within one year
# at exact age x. Source: https://www.ssa.gov/oact/STATS/table4c6.html
# Age 100 is set to 1.0 (terminal age for this model).
SSA_LIFE_TABLE_QX_MALE: dict[int, float] = {
    62: 0.012457, 63: 0.013526, 64: 0.014752, 65: 0.016049, 66: 0.017523,
    67: 0.019136, 68: 0.020949, 69: 0.022977, 70: 0.025261, 71: 0.027850,
    72: 0.030649, 73: 0.033894, 74: 0.037549, 75: 0.041699, 76: 0.046447,
    77: 0.051735, 78: 0.057699, 79: 0.064465, 80: 0.072223, 81: 0.081089,
    82: 0.091104, 83: 0.102310, 84: 0.114779, 85: 0.128592, 86: 0.143818,
    87: 0.160513, 88: 0.178710, 89: 0.198415, 90: 0.219595, 91: 0.242173,
    92: 0.266016, 93: 0.290928, 94: 0.316645, 95: 0.342822, 96: 0.369031,
    97: 0.394760, 98: 0.419430, 99: 0.442419, 100: 1.000000,
}
SSA_LIFE_TABLE_QX_FEMALE: dict[int, float] = {
    62: 0.007830, 63: 0.008543, 64: 0.009360, 65: 0.010254, 66: 0.011258,
    67: 0.012368, 68: 0.013629, 69: 0.015073, 70: 0.016753, 71: 0.018697,
    72: 0.020868, 73: 0.023417, 74: 0.026303, 75: 0.029646, 76: 0.033532,
    77: 0.037932, 78: 0.043012, 79: 0.048884, 80: 0.055756, 81: 0.063734,
    82: 0.072892, 83: 0.083389, 84: 0.095392, 85: 0.109028, 86: 0.124378,
    87: 0.141460, 88: 0.160249, 89: 0.180649, 90: 0.202501, 91: 0.225581,
    92: 0.249587, 93: 0.274136, 94: 0.298766, 95: 0.322933, 96: 0.346023,
    97: 0.367375, 98: 0.386304, 99: 0.402123, 100: 1.000000,
}


def get_life_table(sex: str) -> dict[int, float]:
    if sex.lower() == "female":
        return SSA_LIFE_TABLE_QX_FEMALE
    return SSA_LIFE_TABLE_QX_MALE


def compute_death_probabilities(
    current_age: int,
    max_age: int,
    qx_table: dict[int, float],
) -> dict[int, float]:
    """Conditional probability of dying during each age year, given alive at
    current_age.

    Returns {age: P(die during age | alive at current_age)} for ages from
    current_age to max_age. Probabilities sum to 1.0 (everyone dies by
    max_age, where q(max_age) is forced to 1.0).
    """

    probs: dict[int, float] = {}
    survival = 1.0
    for age in range(current_age, max_age + 1):
        qx = qx_table.get(age, 1.0)
        probs[age] = survival * qx
        survival *= (1 - qx)
    return probs


def compute_life_expectancy(
    current_age: int,
    max_age: int,
    qx_table: dict[int, float],
) -> float:
    """Remaining life expectancy in years, conditional on being alive at
    current_age."""

    death_probs = compute_death_probabilities(current_age, max_age, qx_table)
    return sum(age * prob for age, prob in death_probs.items()) - current_age


@dataclass(frozen=True)
class Age:
    """Age represented as whole years plus additional months."""

    years: int
    months: int = 0

    @property
    def total_months(self) -> int:
        return self.years * MONTHS_PER_YEAR + self.months


@dataclass(frozen=True)
class EarningsTestLimits:
    """
    Retirement earnings test limits for a calendar year.

    SSA has two exempt amounts:
        - lower_limit: used before the calendar year the worker reaches FRA.
        - higher_limit: used in the calendar year the worker reaches FRA, but
          only for earnings before the month FRA is reached.

    Source: https://www.ssa.gov/OACT/COLA/rtea.html
    """

    lower_limit: float
    higher_limit: float


def parse_iso_date(value: str) -> date:
    return date.fromisoformat(value)


def parse_age(value: Any) -> Age:
    """Parse an age from a number (62, 66.83) or object ({years, months})."""

    if isinstance(value, dict):
        return Age(years=int(value["years"]), months=int(value.get("months", 0)))

    years = int(value)
    months = round((float(value) - years) * MONTHS_PER_YEAR)
    return Age(years=years, months=months)


def add_months(start: date, months: int) -> date:
    month_index = start.month - 1 + months
    year = start.year + month_index // MONTHS_PER_YEAR
    month = month_index % MONTHS_PER_YEAR + 1

    days_in_month = [
        31,
        29 if is_leap_year(year) else 28,
        31,
        30,
        31,
        30,
        31,
        31,
        30,
        31,
        30,
        31,
    ]
    day = min(start.day, days_in_month[month - 1])
    return date(year, month, day)


def is_leap_year(year: int) -> bool:
    return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)


def months_between(start: date, end: date) -> int:
    return (end.year - start.year) * MONTHS_PER_YEAR + (end.month - start.month)


def age_to_label(age: Age) -> str:
    if age.months:
        return f"{age.years}y {age.months}m"
    return str(age.years)


def claiming_multiplier(claim_age: Age, fra_age: Age) -> float:
    """
    Permanent start-age multiplier applied to the FRA benefit.

    Early retirement reductions:
        - 5/9 of 1% per month for the first 36 months early
        - 5/12 of 1% per month for additional months early
    Delayed credits (1943 or later):
        - 2/3 of 1% per month (8% per year), stopping at age 70
    Sources:
        https://www.ssa.gov/benefits/retirement/planner/agereduction.html
        https://www.ssa.gov/OACT/ProgData/ar_drc.html
    """

    claim_months = claim_age.total_months
    fra_months = fra_age.total_months

    if claim_months < fra_months:
        months_early = fra_months - claim_months
        first_36 = min(months_early, 36)
        additional = max(months_early - 36, 0)
        reduction = first_36 * (5 / 9 / 100) + additional * (5 / 12 / 100)
        return 1 - reduction

    months_delayed = min(claim_months - fra_months, 70 * MONTHS_PER_YEAR - fra_months)
    delayed_credit = months_delayed * (2 / 3 / 100)
    return 1 + delayed_credit


def multiplier_after_withheld_month_credit(
    original_claim_age: Age,
    fra_age: Age,
    credited_months: float,
) -> float:
    """
    Recalculate the post-FRA multiplier after work-related withholding.

    Simplifying assumption: convert total withheld dollars into equivalent
    fully-withheld months, then reduce the original early-claiming month count
    by that amount.
    """

    original_months_early = max(fra_age.total_months - original_claim_age.total_months, 0)
    adjusted_months_early = max(original_months_early - credited_months, 0)

    first_36 = min(adjusted_months_early, 36)
    additional = max(adjusted_months_early - 36, 0)
    reduction = first_36 * (5 / 9 / 100) + additional * (5 / 12 / 100)
    return 1 - reduction


def load_earnings_test_limits(raw_limits: dict[str, Any]) -> dict[int, EarningsTestLimits]:
    return {
        int(year): EarningsTestLimits(
            lower_limit=float(values["lower_limit"]),
            higher_limit=float(values["higher_limit"]),
        )
        for year, values in raw_limits.items()
    }


def get_limits_for_year(
    year: int,
    limits_by_year: dict[int, EarningsTestLimits],
    default_limits: EarningsTestLimits,
) -> EarningsTestLimits:
    return limits_by_year.get(year, default_limits)


def resolve_salary_for_age(salary_config: Any, age_years: int) -> float:
    """
    Resolve annual salary for a person of `age_years`.

    Accepts:
        - number: same salary for every age.
        - object {age_string: salary}: last-observation carried forward. Ages
          below the earliest listed key are treated as $0 (assumes not yet
          earning at that age); ages above the latest key inherit the latest
          value.
    """

    if isinstance(salary_config, (int, float)):
        return float(salary_config)
    if isinstance(salary_config, dict):
        by_age = sorted((int(k), float(v)) for k, v in salary_config.items())
        applicable = [v for a, v in by_age if a <= age_years]
        if not applicable:
            return 0.0
        return applicable[-1]
    raise ValueError(
        "annual_salary must be a number or an object mapping age to salary; "
        f"got {type(salary_config).__name__}"
    )


def resolve_starting_monthly_benefit(
    config: dict[str, Any],
    claim_age: Age,
    fra_age: Age,
    fra_monthly_benefit: float,
) -> tuple[float, str]:
    """
    Return (starting_monthly_benefit, source_label).

    If `benefits_by_claim_age` is provided in the config and contains an entry
    matching this claim age's whole-year value (with months == 0), use that
    number verbatim. Otherwise derive from PIA using SSA's percentage rules.
    """

    overrides = config.get("benefits_by_claim_age") or {}
    if overrides and claim_age.months == 0:
        by_age = {int(k): float(v) for k, v in overrides.items()}
        if claim_age.years in by_age:
            return by_age[claim_age.years], "ssa_statement_override"

    multiplier = claiming_multiplier(claim_age, fra_age)
    return fra_monthly_benefit * multiplier, "derived_from_pia"


def calculate_calendar_year_withholding(
    *,
    year: int,
    annual_salary: float,
    gross_by_month: dict[int, float],
    fra_date: date,
    limits: EarningsTestLimits,
) -> dict[int, float]:
    """Allocate retirement earnings-test withholding across months in a year."""

    withholding_by_month = {month: 0.0 for month in gross_by_month}

    benefit_months_before_fra = [
        month
        for month in gross_by_month
        if date(year, month, 1) < date(fra_date.year, fra_date.month, 1)
    ]
    if not benefit_months_before_fra:
        return withholding_by_month

    monthly_salary = annual_salary / MONTHS_PER_YEAR

    if year < fra_date.year:
        excess_earnings = max(annual_salary - limits.lower_limit, 0)
        annual_withholding = excess_earnings / 2
    elif year == fra_date.year:
        months_before_fra = fra_date.month - 1
        earnings_before_fra_month = monthly_salary * months_before_fra
        excess_earnings = max(earnings_before_fra_month - limits.higher_limit, 0)
        annual_withholding = excess_earnings / 3
    else:
        annual_withholding = 0

    remaining = min(
        annual_withholding,
        sum(gross_by_month[month] for month in benefit_months_before_fra),
    )

    for month in sorted(benefit_months_before_fra):
        month_gross = gross_by_month[month]
        month_withheld = min(month_gross, remaining)
        withholding_by_month[month] = month_withheld
        remaining -= month_withheld
        if remaining <= 0:
            break

    return withholding_by_month


def project_claim_scenario(
    config: dict[str, Any],
    claim_age: Age,
    horizon_age: Age | None = None,
) -> list[dict[str, Any]]:
    """Build age-year rows for one claiming-age scenario.

    `horizon_age` optionally overrides the projection end age (used to extend
    projections past the expected end-of-life age for the sensitivity table).
    """

    birth_date = parse_iso_date(config["birth_date"])
    fra_age = parse_age(config["full_retirement_age"])
    end_of_life_age = horizon_age if horizon_age is not None else parse_age(config["end_of_life_age"])
    salary_config = config["annual_salary"]
    fra_monthly_benefit = float(config["ssa_full_retirement_monthly_benefit"])
    adjust_at_fra = bool(config.get("adjust_benefit_at_fra_for_deferred_months", True))

    default_limits = EarningsTestLimits(
        lower_limit=float(config["default_earnings_test_limits"]["lower_limit"]),
        higher_limit=float(config["default_earnings_test_limits"]["higher_limit"]),
    )
    limits_by_year = load_earnings_test_limits(config.get("earnings_test_limits_by_year", {}))

    claim_date = add_months(birth_date, claim_age.total_months)
    fra_date = add_months(birth_date, fra_age.total_months)
    end_date = add_months(birth_date, end_of_life_age.total_months)

    starting_monthly_benefit, benefit_source = resolve_starting_monthly_benefit(
        config, claim_age, fra_age, fra_monthly_benefit
    )

    total_months = months_between(claim_date, end_date)
    monthly_records: list[dict[str, Any]] = []

    for offset in range(total_months):
        current_date = add_months(claim_date, offset)
        attained_age_months = months_between(birth_date, current_date)
        monthly_records.append(
            {
                "date": current_date,
                "attained_age": attained_age_months // MONTHS_PER_YEAR,
                "gross_benefit": starting_monthly_benefit,
                "withheld_due_to_work": 0.0,
            }
        )

    records_by_year: dict[int, list[dict[str, Any]]] = {}
    for record in monthly_records:
        records_by_year.setdefault(record["date"].year, []).append(record)

    total_withheld_before_fra = 0.0
    for year, records in records_by_year.items():
        gross_by_month = {record["date"].month: record["gross_benefit"] for record in records}
        limits = get_limits_for_year(year, limits_by_year, default_limits)

        attained_age_at_year_start = months_between(birth_date, date(year, 1, 1)) // MONTHS_PER_YEAR
        annual_salary_this_year = resolve_salary_for_age(salary_config, attained_age_at_year_start)

        withholding_by_month = calculate_calendar_year_withholding(
            year=year,
            annual_salary=annual_salary_this_year,
            gross_by_month=gross_by_month,
            fra_date=fra_date,
            limits=limits,
        )

        for record in records:
            withheld = withholding_by_month[record["date"].month]
            record["withheld_due_to_work"] = withheld
            if record["date"] < fra_date:
                total_withheld_before_fra += withheld

    post_fra_monthly_benefit = starting_monthly_benefit
    credited_months = 0.0
    if (
        adjust_at_fra
        and claim_age.total_months < fra_age.total_months
        and starting_monthly_benefit > 0
    ):
        credited_months = total_withheld_before_fra / starting_monthly_benefit
        adjusted_multiplier = multiplier_after_withheld_month_credit(
            original_claim_age=claim_age,
            fra_age=fra_age,
            credited_months=credited_months,
        )
        post_fra_monthly_benefit = fra_monthly_benefit * adjusted_multiplier

        for record in monthly_records:
            if record["date"] >= fra_date:
                record["gross_benefit"] = post_fra_monthly_benefit

    rows_by_age: dict[int, dict[str, Any]] = {}
    for record in monthly_records:
        age = record["attained_age"]
        gross = record["gross_benefit"]
        withheld = record["withheld_due_to_work"]
        paid = gross - withheld

        row = rows_by_age.setdefault(
            age,
            {
                "claim_age": age_to_label(claim_age),
                "age": age,
                "salary": resolve_salary_for_age(salary_config, age),
                "months_paid_in_age_year": 0,
                "monthly_benefit_at_year_end": gross,
                "gross_social_security": 0.0,
                "withheld_deferred_due_to_work": 0.0,
                "paid_social_security": 0.0,
                "credited_months_at_fra": credited_months,
                "benefit_source": benefit_source,
                "notes": "",
            },
        )
        row["months_paid_in_age_year"] += 1
        row["monthly_benefit_at_year_end"] = gross
        row["gross_social_security"] += gross
        row["withheld_deferred_due_to_work"] += withheld
        row["paid_social_security"] += paid

    cumulative_paid = 0.0
    cumulative_deferred = 0.0
    output_rows = []
    for age in sorted(rows_by_age):
        row = rows_by_age[age]
        cumulative_paid += row["paid_social_security"]
        cumulative_deferred += row["withheld_deferred_due_to_work"]
        row["cumulative_paid_social_security"] = cumulative_paid
        row["cumulative_deferred_due_to_work"] = cumulative_deferred

        if row["withheld_deferred_due_to_work"] > 0:
            row["notes"] = "Earnings-test withholding tracked as deferred, not lost."
        elif age >= fra_age.years:
            row["notes"] = "No earnings-test withholding at or after FRA month."

        output_rows.append(round_money_fields(row))

    return output_rows


def round_money_fields(row: dict[str, Any]) -> dict[str, Any]:
    money_fields = {
        "salary",
        "monthly_benefit_at_year_end",
        "gross_social_security",
        "withheld_deferred_due_to_work",
        "paid_social_security",
        "cumulative_paid_social_security",
        "cumulative_deferred_due_to_work",
    }
    rounded = dict(row)
    for field in money_fields:
        rounded[field] = round(float(rounded[field]), 2)
    rounded["credited_months_at_fra"] = round(float(rounded["credited_months_at_fra"]), 3)
    return rounded


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return

    fieldnames = list(rows[0].keys())
    with path.open("w", newline="") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


# ---------------------------------------------------------------------------
# Break-even analysis
# ---------------------------------------------------------------------------


def cumulative_paid_at_age(rows: list[dict[str, Any]], age: int) -> float:
    """Cumulative paid Social Security through the end of `age`.

    Returns 0 if `age` is before the scenario's claim age.
    Returns the scenario's final cumulative if `age` is beyond the last row.
    """

    result = 0.0
    for row in rows:
        if row["age"] <= age:
            result = row["cumulative_paid_social_security"]
        else:
            break
    return result


def find_break_even_age(
    earlier_rows: list[dict[str, Any]],
    later_rows: list[dict[str, Any]],
    min_age: int,
    max_age: int,
) -> int | None:
    """First whole-year age at which the later scenario's cumulative paid
    meets or exceeds the earlier scenario's cumulative paid.

    Returns None if the later scenario never catches up within [min_age, max_age].
    """

    for age in range(min_age, max_age + 1):
        earlier_cum = cumulative_paid_at_age(earlier_rows, age)
        later_cum = cumulative_paid_at_age(later_rows, age)
        if later_cum >= earlier_cum and later_cum > 0:
            return age
    return None


def compute_break_even_matrix_and_details(
    scenarios: list[tuple[Age, list[dict[str, Any]]]],
    min_age: int,
    max_age: int,
    expected_lifespan_age: int,
) -> tuple[list[list[Any]], list[dict[str, Any]]]:
    """
    Returns (matrix_2d, details_rows).

    matrix_2d:
        Compact grid. Row = later claim age, column = earlier claim age.
        Cell value = break-even whole-year age (or ">MAX" if it never catches
        up within the horizon).

    details_rows:
        One row per ordered pair (earlier, later) with cumulative-paid values
        at expected lifespan and a plain-English verdict.
    """

    scenarios_sorted = sorted(scenarios, key=lambda x: x[0].total_months)
    labels = [age_to_label(age) for age, _ in scenarios_sorted]

    header = ["Later \\ Earlier"] + [f"Claim {lbl}" for lbl in labels]
    matrix: list[list[Any]] = [header]

    details: list[dict[str, Any]] = []

    for i, (later_age, later_rows) in enumerate(scenarios_sorted):
        row_cells: list[Any] = [f"Claim {labels[i]}"]
        for j, (earlier_age, earlier_rows) in enumerate(scenarios_sorted):
            if j >= i:
                row_cells.append("—")
                continue

            be_age = find_break_even_age(earlier_rows, later_rows, min_age, max_age)
            row_cells.append(be_age if be_age is not None else f">{max_age}")

            cum_earlier = cumulative_paid_at_age(earlier_rows, expected_lifespan_age)
            cum_later = cumulative_paid_at_age(later_rows, expected_lifespan_age)
            lifetime_diff = cum_later - cum_earlier

            if be_age is None:
                verdict = (
                    f"Claim at {labels[j]} — waiting to {labels[i]} never catches up "
                    f"before age {max_age}."
                )
            elif be_age <= expected_lifespan_age:
                verdict = (
                    f"Wait until {labels[i]} — breaks even at age {be_age}, "
                    f"{expected_lifespan_age - be_age} years before your expected lifespan."
                )
            else:
                verdict = (
                    f"Claim at {labels[j]} — waiting to {labels[i]} only breaks even at "
                    f"age {be_age}, {be_age - expected_lifespan_age} years after your "
                    f"expected lifespan."
                )

            details.append(
                {
                    "earlier_claim_age": labels[j],
                    "later_claim_age": labels[i],
                    "break_even_age": be_age if be_age is not None else f">{max_age}",
                    "cumulative_paid_earlier_at_expected_lifespan": round(cum_earlier, 2),
                    "cumulative_paid_later_at_expected_lifespan": round(cum_later, 2),
                    "lifetime_difference_later_minus_earlier": round(lifetime_diff, 2),
                    "verdict_at_expected_lifespan": verdict,
                }
            )
        matrix.append(row_cells)

    return matrix, details


def compute_lifespan_sensitivity(
    scenarios: list[tuple[Age, list[dict[str, Any]]]],
    checkpoint_ages: list[int],
    expected_lifespan_age: int,
) -> tuple[list[str], list[dict[str, Any]]]:
    """
    Cumulative paid Social Security for each scenario at each checkpoint age.

    Returns (checkpoint_labels_in_order, rows). Rows are ordered by claim age.
    """

    checkpoints = sorted(set(list(checkpoint_ages) + [expected_lifespan_age]))
    labels = [
        f"cum_paid_at_age_{age}" + (" (expected_lifespan)" if age == expected_lifespan_age else "")
        for age in checkpoints
    ]

    rows: list[dict[str, Any]] = []
    for age, scenario_rows in sorted(scenarios, key=lambda x: x[0].total_months):
        row: dict[str, Any] = {"claim_age": age_to_label(age)}
        for chk_age, chk_label in zip(checkpoints, labels):
            row[chk_label] = round(cumulative_paid_at_age(scenario_rows, chk_age), 2)
        rows.append(row)

    return labels, rows


def compute_expected_value_analysis(
    scenarios: list[tuple[Age, list[dict[str, Any]]]],
    current_age: int,
    qx_table: dict[int, float],
    max_age: int = ACTUARIAL_MAX_AGE,
) -> list[dict[str, Any]]:
    """For each claim-age scenario, compute the actuarial expected lifetime
    benefit weighted by mortality probabilities.

    Returns rows sorted by expected benefit (best first), each containing:
        claim_age, expected_benefit, life_expectancy,
        P(survive to 75/80/85/90), cumulative_at_life_expectancy
    """

    death_probs = compute_death_probabilities(current_age, max_age, qx_table)
    life_exp = compute_life_expectancy(current_age, max_age, qx_table)

    survival_checkpoints = [75, 80, 85, 90]
    survival_probs: dict[int, float] = {}
    for chk in survival_checkpoints:
        survival_probs[chk] = 1.0 - sum(
            p for age, p in death_probs.items() if age < chk
        )

    results: list[dict[str, Any]] = []
    for claim_age, scenario_rows in sorted(scenarios, key=lambda x: x[0].total_months):
        expected_benefit = sum(
            prob * cumulative_paid_at_age(scenario_rows, death_age)
            for death_age, prob in death_probs.items()
        )
        cum_at_life_exp = cumulative_paid_at_age(
            scenario_rows, round(current_age + life_exp)
        )

        row: dict[str, Any] = {
            "claim_age": age_to_label(claim_age),
            "expected_lifetime_benefit": round(expected_benefit, 2),
            "life_expectancy_from_current_age": round(life_exp, 1),
            "cumulative_at_life_expectancy": round(cum_at_life_exp, 2),
        }
        for chk in survival_checkpoints:
            row[f"P_survive_to_{chk}"] = round(survival_probs[chk] * 100, 1)

        results.append(row)

    results.sort(key=lambda r: r["expected_lifetime_benefit"], reverse=True)
    return results


def build_summary(
    config: dict[str, Any],
    scenarios: list[tuple[Age, list[dict[str, Any]]]],
    end_of_life_age: Age,
    ev_results: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Build the summary payload rendered on the Excel Summary tab and to stdout."""

    sorted_scenarios = sorted(scenarios, key=lambda x: x[0].total_months)
    expected = end_of_life_age.years

    best_age, best_rows = max(
        sorted_scenarios,
        key=lambda s: cumulative_paid_at_age(s[1], expected),
    )
    best_amount = cumulative_paid_at_age(best_rows, expected)

    consecutive_verdicts: list[str] = []
    for i in range(len(sorted_scenarios) - 1):
        earlier_age, earlier_rows = sorted_scenarios[i]
        later_age, later_rows = sorted_scenarios[i + 1]
        be_age = find_break_even_age(earlier_rows, later_rows, 62, expected)
        e_lbl = age_to_label(earlier_age)
        l_lbl = age_to_label(later_age)

        if be_age is None:
            consecutive_verdicts.append(
                f"{e_lbl} vs {l_lbl}: waiting never catches up by age {expected}. "
                f"CLAIM AT {e_lbl}."
            )
        elif be_age <= expected:
            consecutive_verdicts.append(
                f"{e_lbl} vs {l_lbl}: breaks even at age {be_age} — "
                f"{expected - be_age} years before your expected lifespan. WAIT to {l_lbl}."
            )
        else:
            consecutive_verdicts.append(
                f"{e_lbl} vs {l_lbl}: breaks even at age {be_age} — "
                f"{be_age - expected} years AFTER your expected lifespan. CLAIM AT {e_lbl}."
            )

    salary_config = config["annual_salary"]
    if isinstance(salary_config, (int, float)):
        salary_display = f"${float(salary_config):,.0f} / year (flat)"
    else:
        parts = [f"age {k}: ${float(v):,.0f}" for k, v in sorted(salary_config.items(), key=lambda x: int(x[0]))]
        salary_display = "; ".join(parts) + " (last value carried forward)"

    return {
        "birth_date": config["birth_date"],
        "full_retirement_age": age_to_label(parse_age(config["full_retirement_age"])),
        "expected_end_of_life_age": age_to_label(end_of_life_age),
        "fra_monthly_benefit_pia": float(config["ssa_full_retirement_monthly_benefit"]),
        "salary_input": salary_display,
        "benefits_source": (
            "SSA-statement overrides where available; PIA-derived otherwise"
            if config.get("benefits_by_claim_age")
            else "PIA-derived from FRA monthly benefit"
        ),
        "best_strategy_at_expected_lifespan": {
            "claim_age": age_to_label(best_age),
            "cumulative_paid": round(best_amount, 2),
        },
        "consecutive_pair_verdicts": consecutive_verdicts,
        "model_caveats": [
            "No COLA, no inflation, no taxation, no Medicare premiums.",
            "Salary uses last-observation-carried-forward when given as an age map.",
            "FRA recomputation uses PIA-based math even when SSA-statement overrides are supplied.",
            "Benefits withheld by the earnings test are counted as deferred, not lost.",
            "Break-even is cumulative cash received (no NPV / discount rate).",
            "Expected-value uses SSA 2021 Period Life Table (period, not cohort).",
        ],
        "expected_value": ev_results,
    }


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------


MONEY_FORMAT = "$#,##0"
MONEY_FORMAT_PRECISE = "$#,##0.00"


def write_excel(
    path: Path,
    projection_rows: list[dict[str, Any]],
    matrix: list[list[Any]],
    break_even_details: list[dict[str, Any]],
    sensitivity_labels: list[str],
    sensitivity_rows: list[dict[str, Any]],
    summary: dict[str, Any],
    expected_lifespan_age: int,
    ev_results: list[dict[str, Any]] | None = None,
) -> bool:
    """Write the multi-tab workbook. Returns True on success, False if openpyxl
    is unavailable."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _write_summary_sheet(ws, summary)

    ws = wb.create_sheet("Break-Even")
    _write_break_even_sheet(ws, matrix, break_even_details, expected_lifespan_age)

    ws = wb.create_sheet("Sensitivity")
    _write_sensitivity_sheet(ws, sensitivity_labels, sensitivity_rows, expected_lifespan_age)

    if ev_results:
        ws = wb.create_sheet("Expected Value")
        _write_expected_value_sheet(ws, ev_results)

    ws = wb.create_sheet("Projection")
    _write_projection_sheet(ws, projection_rows)

    wb.save(path)
    return True


def _write_summary_sheet(ws, summary: dict[str, Any]) -> None:
    from openpyxl.styles import Font, PatternFill

    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    section_fill = PatternFill("solid", fgColor="E8EEF7")

    ws["A1"] = "Social Security Break-Even Analysis"
    ws["A1"].font = Font(bold=True, size=16)

    row = 3

    def section(title: str):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        row += 1

    section("Inputs")
    for label, value in [
        ("Birth date", summary["birth_date"]),
        ("Full retirement age (FRA)", summary["full_retirement_age"]),
        ("Expected end-of-life age", summary["expected_end_of_life_age"]),
        ("FRA monthly benefit (PIA)", summary["fra_monthly_benefit_pia"]),
        ("Salary input", summary["salary_input"]),
        ("Benefits source", summary["benefits_source"]),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if label == "FRA monthly benefit (PIA)":
            ws.cell(row=row, column=2).number_format = MONEY_FORMAT_PRECISE
        row += 1

    row += 1
    section("Best strategy at expected lifespan")
    best = summary["best_strategy_at_expected_lifespan"]
    ws.cell(row=row, column=1, value="Claim at age").font = Font(bold=True)
    ws.cell(row=row, column=2, value=best["claim_age"]).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Cumulative paid at expected lifespan")
    cell = ws.cell(row=row, column=2, value=best["cumulative_paid"])
    cell.number_format = MONEY_FORMAT
    row += 2

    section("Consecutive-pair recommendations")
    for verdict in summary["consecutive_pair_verdicts"]:
        ws.cell(row=row, column=1, value=verdict)
        row += 1

    ev_results = summary.get("expected_value")
    if ev_results:
        row += 1
        section("Actuarial expected-value analysis (SSA 2021 Life Table)")
        best_ev = ev_results[0]
        ws.cell(row=row, column=1, value="Best strategy (mortality-weighted)").font = Font(bold=True)
        ws.cell(row=row, column=2, value=best_ev["claim_age"]).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Expected lifetime benefit")
        cell = ws.cell(row=row, column=2, value=best_ev["expected_lifetime_benefit"])
        cell.number_format = MONEY_FORMAT
        row += 1
        ws.cell(row=row, column=1, value="Life expectancy from current age")
        ws.cell(row=row, column=2, value=f"{best_ev['life_expectancy_from_current_age']} years")
        row += 1
        ws.cell(row=row, column=1, value="")
        row += 1
        for ev_row in ev_results:
            ws.cell(row=row, column=1, value=f"Claim {ev_row['claim_age']}")
            cell = ws.cell(row=row, column=2, value=ev_row["expected_lifetime_benefit"])
            cell.number_format = MONEY_FORMAT
            row += 1

    row += 1
    section("Model caveats")
    for caveat in summary["model_caveats"]:
        ws.cell(row=row, column=1, value=f"• {caveat}")
        row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 60


def _write_break_even_sheet(
    ws, matrix: list[list[Any]], details: list[dict[str, Any]], expected_lifespan_age: int
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    good_fill = PatternFill("solid", fgColor="D6EFD6")
    bad_fill = PatternFill("solid", fgColor="F7D6D6")

    ws["A1"] = "Break-Even Matrix — first age at which waiting catches up"
    ws["A1"].font = Font(bold=True, size=13)

    ws.cell(
        row=2,
        column=1,
        value=(
            f"Green = breaks even at or before expected lifespan ({expected_lifespan_age}). "
            f"Red = breaks even after expected lifespan or never."
        ),
    )

    matrix_start_row = 4
    for r, row_cells in enumerate(matrix):
        for c, cell_value in enumerate(row_cells):
            cell = ws.cell(row=matrix_start_row + r, column=1 + c, value=cell_value)
            if r == 0 or c == 0:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            else:
                if isinstance(cell_value, int):
                    if cell_value <= expected_lifespan_age:
                        cell.fill = good_fill
                    else:
                        cell.fill = bad_fill
                elif isinstance(cell_value, str) and cell_value.startswith(">"):
                    cell.fill = bad_fill
                cell.alignment = Alignment(horizontal="center")

    for c in range(1, len(matrix[0]) + 1):
        ws.column_dimensions[chr(64 + c) if c <= 26 else "AA"].width = 16

    details_start_row = matrix_start_row + len(matrix) + 3
    ws.cell(row=details_start_row - 1, column=1, value="Pairwise details").font = Font(
        bold=True, size=13
    )

    if not details:
        return

    fieldnames = list(details[0].keys())
    for c, name in enumerate(fieldnames):
        cell = ws.cell(row=details_start_row, column=1 + c, value=name)
        cell.font = header_font
        cell.fill = header_fill

    money_columns = {
        "cumulative_paid_earlier_at_expected_lifespan",
        "cumulative_paid_later_at_expected_lifespan",
        "lifetime_difference_later_minus_earlier",
    }
    for r, row in enumerate(details):
        for c, name in enumerate(fieldnames):
            cell = ws.cell(row=details_start_row + 1 + r, column=1 + c, value=row[name])
            if name in money_columns:
                cell.number_format = MONEY_FORMAT

    for c, name in enumerate(fieldnames):
        col_letter = chr(64 + 1 + c) if 1 + c <= 26 else "AA"
        ws.column_dimensions[col_letter].width = max(20, min(50, len(name) + 4))


def _write_sensitivity_sheet(
    ws,
    labels: list[str],
    rows: list[dict[str, Any]],
    expected_lifespan_age: int,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    expected_fill = PatternFill("solid", fgColor="FFF3B8")

    ws["A1"] = f"Lifespan Sensitivity — cumulative Social Security paid, by age"
    ws["A1"].font = Font(bold=True, size=13)

    fieldnames = ["claim_age"] + labels
    for c, name in enumerate(fieldnames):
        cell = ws.cell(row=3, column=1 + c, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows):
        ws.cell(row=4 + r, column=1, value=row["claim_age"]).font = Font(bold=True)
        for c, label in enumerate(labels, start=1):
            cell = ws.cell(row=4 + r, column=1 + c, value=row[label])
            cell.number_format = MONEY_FORMAT
            if "expected_lifespan" in label:
                cell.fill = expected_fill

    ws.column_dimensions["A"].width = 14
    for c in range(len(labels)):
        col_letter = chr(66 + c) if 66 + c <= ord("Z") else "AA"
        ws.column_dimensions[col_letter].width = 26


def _write_expected_value_sheet(ws, ev_results: list[dict[str, Any]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    best_fill = PatternFill("solid", fgColor="D6EFD6")

    ws["A1"] = "Actuarial Expected Value — mortality-weighted lifetime benefit"
    ws["A1"].font = Font(bold=True, size=13)

    ws.cell(
        row=2,
        column=1,
        value=(
            "Each claim age's expected benefit is weighted by the SSA 2021 Period Life Table — "
            "the probability of dying at each future age. No single lifespan assumption needed."
        ),
    )

    if not ev_results:
        return

    fieldnames = list(ev_results[0].keys())
    display_names = {
        "claim_age": "Claim Age",
        "expected_lifetime_benefit": "Expected Lifetime Benefit",
        "life_expectancy_from_current_age": "Life Expectancy (yrs from now)",
        "cumulative_at_life_expectancy": "Cumulative at Life Expectancy",
        "P_survive_to_75": "P(survive to 75) %",
        "P_survive_to_80": "P(survive to 80) %",
        "P_survive_to_85": "P(survive to 85) %",
        "P_survive_to_90": "P(survive to 90) %",
    }

    for c, name in enumerate(fieldnames):
        cell = ws.cell(row=4, column=1 + c, value=display_names.get(name, name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    money_cols = {"expected_lifetime_benefit", "cumulative_at_life_expectancy"}
    pct_cols = {k for k in fieldnames if k.startswith("P_survive")}

    for r, row in enumerate(ev_results):
        for c, name in enumerate(fieldnames):
            cell = ws.cell(row=5 + r, column=1 + c, value=row[name])
            if name in money_cols:
                cell.number_format = MONEY_FORMAT
            elif name in pct_cols:
                cell.number_format = "0.0"
            cell.alignment = Alignment(horizontal="center")
        if r == 0:
            for c in range(len(fieldnames)):
                ws.cell(row=5 + r, column=1 + c).fill = best_fill

    widths = {
        "claim_age": 12,
        "expected_lifetime_benefit": 24,
        "life_expectancy_from_current_age": 28,
        "cumulative_at_life_expectancy": 26,
    }
    for c, name in enumerate(fieldnames):
        col_letter = chr(65 + c) if c < 26 else "AA"
        ws.column_dimensions[col_letter].width = widths.get(name, 20)


def _write_projection_sheet(ws, rows: list[dict[str, Any]]) -> None:
    from openpyxl.styles import Font, PatternFill

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF7")

    if not rows:
        return

    fieldnames = list(rows[0].keys())
    for c, name in enumerate(fieldnames):
        cell = ws.cell(row=1, column=1 + c, value=name)
        cell.font = header_font
        cell.fill = header_fill

    money_fields = {
        "salary",
        "monthly_benefit_at_year_end",
        "gross_social_security",
        "withheld_deferred_due_to_work",
        "paid_social_security",
        "cumulative_paid_social_security",
        "cumulative_deferred_due_to_work",
    }
    for r, row in enumerate(rows):
        for c, name in enumerate(fieldnames):
            cell = ws.cell(row=2 + r, column=1 + c, value=row[name])
            if name in money_fields:
                cell.number_format = MONEY_FORMAT_PRECISE

    widths = {
        "claim_age": 10,
        "age": 6,
        "salary": 14,
        "months_paid_in_age_year": 10,
        "monthly_benefit_at_year_end": 18,
        "gross_social_security": 18,
        "withheld_deferred_due_to_work": 22,
        "paid_social_security": 18,
        "credited_months_at_fra": 14,
        "benefit_source": 26,
        "notes": 60,
        "cumulative_paid_social_security": 24,
        "cumulative_deferred_due_to_work": 24,
    }
    for c, name in enumerate(fieldnames):
        col_letter = chr(65 + c) if c < 26 else "AA"
        ws.column_dimensions[col_letter].width = widths.get(name, 16)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Estimate Social Security benefits while working, with break-even analysis."
    )
    parser.add_argument("input_file", help="Path to a JSON input file.")
    args = parser.parse_args()

    input_path = Path(args.input_file)
    config = json.loads(input_path.read_text())

    claim_ages = [parse_age(value) for value in config["claim_ages"]]
    end_of_life_age = parse_age(config["end_of_life_age"])
    checkpoint_ages = list(config.get("lifespan_checkpoints", DEFAULT_LIFESPAN_CHECKPOINTS))
    sex = config.get("sex", "male")

    projection_horizon = Age(
        years=max(end_of_life_age.years, max(checkpoint_ages, default=0), ACTUARIAL_MAX_AGE)
    )

    scenarios: list[tuple[Age, list[dict[str, Any]]]] = []
    all_rows: list[dict[str, Any]] = []
    for claim_age in claim_ages:
        rows = project_claim_scenario(config, claim_age, horizon_age=projection_horizon)
        scenarios.append((claim_age, rows))
        all_rows.extend(rows)

    output_csv = Path(config.get("output_csv", "social_security_projection.csv"))
    write_csv(output_csv, all_rows)

    matrix, details = compute_break_even_matrix_and_details(
        scenarios,
        min_age=62,
        max_age=end_of_life_age.years,
        expected_lifespan_age=end_of_life_age.years,
    )
    sensitivity_labels, sensitivity_rows = compute_lifespan_sensitivity(
        scenarios,
        checkpoint_ages=checkpoint_ages,
        expected_lifespan_age=end_of_life_age.years,
    )

    birth_date = parse_iso_date(config["birth_date"])
    current_age = (date.today() - birth_date).days // 365
    qx_table = get_life_table(sex)
    ev_results = compute_expected_value_analysis(
        scenarios, current_age, qx_table, max_age=ACTUARIAL_MAX_AGE
    )

    summary = build_summary(config, scenarios, end_of_life_age, ev_results=ev_results)

    output_xlsx = Path(config.get("output_xlsx", "social_security_analysis.xlsx"))
    excel_ok = write_excel(
        output_xlsx,
        projection_rows=all_rows,
        matrix=matrix,
        break_even_details=details,
        sensitivity_labels=sensitivity_labels,
        sensitivity_rows=sensitivity_rows,
        summary=summary,
        expected_lifespan_age=end_of_life_age.years,
        ev_results=ev_results,
    )

    print(f"Wrote {len(all_rows)} projection rows to {output_csv}")
    if excel_ok:
        print(f"Wrote analysis workbook to {output_xlsx}")
    else:
        print(
            "openpyxl is not installed — skipped Excel output. "
            "Install with: pip install openpyxl"
        )
    print()
    print(f"Best strategy at expected lifespan ({age_to_label(end_of_life_age)}): "
          f"claim at {summary['best_strategy_at_expected_lifespan']['claim_age']} "
          f"(${summary['best_strategy_at_expected_lifespan']['cumulative_paid']:,.0f} total)")
    print()
    print("Consecutive-pair verdicts:")
    for verdict in summary["consecutive_pair_verdicts"]:
        print(f"  - {verdict}")

    if ev_results:
        life_exp = ev_results[0]["life_expectancy_from_current_age"]
        print()
        print(f"Actuarial expected value ({sex} life table, "
              f"life expectancy {life_exp} yrs from age {current_age}):")
        for ev_row in ev_results:
            marker = " <-- BEST" if ev_row is ev_results[0] else ""
            print(f"  Claim {ev_row['claim_age']:>5}: "
                  f"${ev_row['expected_lifetime_benefit']:>12,.0f}{marker}")


if __name__ == "__main__":
    main()
